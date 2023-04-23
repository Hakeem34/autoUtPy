#おじどうくんPY
#
#下記で公開されているおじどうくんDR4をPython＋openpyxlに置き換えて実装したものです。
#http://park.ruru.ne.jp/ando/work/autoUt/index_ja.html
#※JAVAの対応は割愛しております


import os
import sys
import re
import datetime
import subprocess
import openpyxl
from openpyxl.styles.borders import Border, Side


programVer  = "ver1.0.0";
programName = "単体試験項目自動作成ツール『おじどうくんPY』";
factor = 1.5
side1      = Side(style='thin', color='000000')
border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)
border_tlr = Border(top=side1, left=side1, right=side1)
border_tb  = Border(top=side1, bottom=side1)
border_lr  = Border(left=side1, right=side1)
border_l   = Border(left=side1)
border_r   = Border(right=side1)
border_t   = Border(top=side1)
border_b   = Border(bottom=side1)
allLen     = 1
pos        = 1

re_direct_ifelse = re.compile(r"\s*\#(if|else)")
re_func_start    = re.compile(r"^([a-zA-Z_][a-zA-Z_0-9]*(\**\s+)?\s+)*(\*(\*|\s)*)?([a-zA-Z][a-zA-Z_0-9]*::)?([a-zA-Z_][a-zA-Z_0-9]*)\s*\(")
re_return        = re.compile(r"^\s+return\s")
re_forwhile      = re.compile(r"^\s+(for|while)(\s|\()")
re_do1           = re.compile(r"^\s+do$")
re_do2           = re.compile(r"^\s+do(\s|\{)")
re_if            = re.compile(r"^\s+if(\s|\()")
re_elseif        = re.compile(r"^\s+(\}\s*)?else\s+if(\s|\()")
re_else          = re.compile(r"^\s+(\}\s*)?else\s+")
re_case          = re.compile(r"^\s+case\s+([^:]+)\s*:")
re_default       = re.compile(r"^\s+default\s+([^:]+)\s*:")
re_switch        = re.compile(r"\s*switch\s*\(")
re_func_call     = re.compile(r"([a-zA-Z][:_a-zA-Z0-9]*)\s*\(")
re_todo          = re.compile(r"TODO")

fill_idx34 = openpyxl.styles.PatternFill(patternType='solid',fgColor='CCFFFF', bgColor='CCFFFF')
fill_idx35 = openpyxl.styles.PatternFill(patternType='solid',fgColor='CCFFCC', bgColor='CCFFCC')
fill_idx36 = openpyxl.styles.PatternFill(patternType='solid',fgColor='FFFF99', bgColor='FFFF99')
wrap_text  = openpyxl.styles.Alignment(wrapText=True)

maxLineNumber = 3000;
text_former = (
"", 
"",
"関数",
"処理のreturnにおいて、戻り値が正しく返ること",
"",
"if(真)の場合にこの処理が起動されること",
"if(偽/else)の場合にこの処理が起動され*ない*こと",
"else if(真)の場合にこの処理が起動されること",
"else の場合にこの処理が起動されること",
"switch()内のパラメタの妥当性を考慮すること",
"case ",
"default: を発生させこの処理が起動されること",
"関数",
"Method ",
"",
"line数が",
)

text_latter = (
" ファイルの試験を実施する",
" の妥当性を検討すること(必要なければ削除する)",
" が呼ばれたときに引数が正しく渡されていること\n(パラメタ数、パラメタ順序、実体orポインタ)\n(引数に最大値最小値を与えてそれぞれチェックすること)",
"",
" が正しい範囲でループされること",
"",
"",
"",
"",
"",
": を発生させこの処理が起動されること",
"",
"において、関数呼び出し時に引数を正しく渡していること\n(パラメータ数、パラメータ順序、実体orポインタ)\n(戻り値が正しく処理されていること)",
" がコールされることを確認すること\n(戻り値が正しく処理されていること)",
"行にある TODO 項目の内容を確認すること",
"行と多いので、分割することを検討すること",
)



#
# ヘルプを出力して終了
def helpExit():
    print("usage:");
    print("ut4.py [OPTION] filename.c [filename.c ...]")
    print("")
    print("[OPTION]")
    print("\t-help : Help message (this).")
    exit(0)



def create_book():
    wb = openpyxl.Workbook()
    return wb


def create_sheet(wb, name):
    global allLen

    fill_idx33 = openpyxl.styles.PatternFill(patternType='solid',fgColor='00CCFF', bgColor='00CCFF')

    ws = wb.create_sheet(name)
    ws.title = name
    ws.cell(1,1).value = programName + "/" + programVer
    allLen += 1
    ws.column_dimensions['A'].width = 14.50 * factor
    ws.column_dimensions['B'].width =  3.88 * factor
    ws.column_dimensions['C'].width =  3.88 * factor
    ws.column_dimensions['D'].width = 16.75 * factor
    ws.column_dimensions['E'].width =  3.88 * factor
    ws.column_dimensions['F'].width =  6.00 * factor
    ws.column_dimensions['G'].width = 50.00 * factor

    ws.cell(2,1).value = "フォルダ名"
    ws.cell(2,2).value = "項番"
    ws.cell(2,3).value = "項番"
    ws.cell(2,4).value = "関数名(ファイル名)"
    ws.cell(2,5).value = "項番"
    ws.cell(2,6).value = "Line"
    ws.cell(2,7).value = "試験項目"

    for col in range(1, 8):
        ws.cell(2, col).fill   = fill_idx33
        ws.cell(2, col).border = border_aro
    allLen += 1
    return ws


def pcall(ws, pattern, file_num, fl, fn, func_num, func_name, pos_arg, line, text):
    global allLen
    global pos

    arg_array = (file_num, fl, fn, func_num, func_name, pos_arg, line, text)

#   print ("pattern : %d" % pattern)
#   print ("args    : %s, %s, %s, %s, %s, %s, %s, %s" % arg_array)
    if (pattern == 0):
        ws.cell(allLen, 1).value = fl
        ws.cell(allLen, 2).value = file_num
        ws.cell(allLen, 3).value = func_num
        ws.cell(allLen, 4).value = fn
    elif (pattern == 2):
        ws.cell(allLen, 3).value = func_num
        ws.cell(allLen, 4).value = func_name

    if (pattern != 15):
        ws.cell(allLen, 5).value = pos_arg

    ws.cell(allLen, 6).value = line
    ws.cell(allLen, 7).value = text_former[pattern] + text + text_latter[pattern]

    for col in range(1, 8):
        if ((ws.cell(allLen, col).value != None) and (ws.cell(allLen, col).value != "")):
            ws.cell(allLen, col).border = border_tlr
        else:
            ws.cell(allLen, col).border = border_lr

    ws.cell(allLen, 1).fill = fill_idx35
    ws.cell(allLen, 2).fill = fill_idx36
    ws.cell(allLen, 3).fill = fill_idx34
    ws.cell(allLen, 4).fill = fill_idx34
    ws.cell(allLen, 1).alignment  = wrap_text
    ws.cell(allLen, 4).alignment  = wrap_text
    ws.cell(allLen, 7).alignment  = wrap_text

    if ((pattern == 2) or (pattern == 12)):
        ws.row_dimensions[allLen].height = 52
    elif (pattern == 13):
        ws.row_dimensions[allLen].height = 26


    allLen += 1
    pos    += 1



def work1Line(ws, funcNum, funcName, line, work):
    global pos

    if (result := re_direct_ifelse.match(work)):
        pcall(ws, 1, "","","",funcNum,funcName,pos,line, work)
    elif (result := re_func_start.match(work)):
        pos = 1
        funcNum += 1
        funcName = result.group(6) + "()";
        pcall(ws, 2, "","","",funcNum,funcName,pos,line, work)
    elif (result := re_return.match(work)):
        pcall(ws, 3, "","","",funcNum,funcName,pos,line, work)
    elif (result := re_forwhile.match(work)):
        selectName = result.group(1) + "()"
        pcall(ws, 4, "","","",funcNum,funcName,pos,line, selectName)
    elif (result := re_do1.match(work)):
        pcall(ws, 4, "","","",funcNum,funcName,pos,line, "do")
    elif (result := re_do2.match(work)):
        pcall(ws, 4, "","","",funcNum,funcName,pos,line, "do")
    elif (result := re_if.match(work)):
        pcall(ws, 5, "","","",funcNum,funcName,pos,line, "if")
        pcall(ws, 6, "","","",funcNum,funcName,pos,line, "if")
    elif (result := re_elseif.match(work)):
        pcall(ws, 7, "","","",funcNum,funcName,pos,line, "else if")
    elif (result := re_else.match(work)):
        pcall(ws, 8, "","","",funcNum,funcName,pos,line, "else")
    elif (result := re_case.match(work)):
        caseName = result.group(1)
        pcall(ws, 9, "","","",funcNum,funcName,pos,line, caseName)
    elif (result := re_default.match(work)):
        pcall(ws, 10, "","","",funcNum,funcName,pos,line, "default")
    elif (result := re_switch.match(work)):
        pcall(ws, 11, "","","",funcNum,funcName,pos,line, "switch")
    elif (result := re_func_call.match(work)):
        functionName = result.group(1) + "()"
        result = re.match(r"^\s*[\/]?\*", work)
        if (result == None):
            pcall(ws, 12, "","","",funcNum,funcName,pos,line, functionName)

    if (result := re_todo.match(work)):
        pcall(ws, 14, "","","",funcNum,funcName,pos,line, functionName)

    return (funcNum, funcName)


#/*****************************************************************************/
#/* メイン関数                                                                */
#/*****************************************************************************/
def main():
    global maxLineNumber

    argc = len(sys.argv)
    if ((argc < 2) or re.match(r"\-+help", sys.argv[1])):
        helpExit()


    wb = create_book()
    
    sys.argv.pop(0)
    fileNumber = 0
    oldFolder  = ""
    for arg in sys.argv:
        fileNumber += 1
        print("open:" + arg )
        file = arg
        fn   = os.path.basename(file)
        fl   = os.path.dirname(file)
        if (fl == ""):
            fl = "."

        if (fl != oldFolder):
            oldFolder = fl
        else:
            fl = ""

        pos = 1
        line = 0
        funcNum = 1
        funcName = ""
        workout = ""

        ws = create_sheet(wb, fn)
        fh = open(file, 'r', encoding="utf-8")
        pcall(ws, 0,fileNumber,fl,fn,funcNum,funcName,pos,line,file)
        read_lines = fh.readlines()
        for line_text in read_lines:
            line += 1
            pattern = r"\/\[\s\S]*$"
            line_text = re.sub(pattern, "", line_text)
            pattern = r"\s+$"
            line_text = re.sub(pattern, "", line_text)
            pattern = r"\/\/.*$"
            line_text = re.sub(pattern, "", line_text)
            pattern = r"\/\*.*\*\/"
            line_text = re.sub(pattern, "", line_text)

            if (result := re.match(r"^[a-zA-Z_][^;\(\{]", line_text)):
                workout = workout + line_text
            else:
                if (workout != ""):
                    (funcNum,funcName) = work1Line(ws, funcNum,funcName,line -1,workout + line_text)
                    workout = "";
                else:
                    (funcNum,funcName) = work1Line(ws, funcNum,funcName,line,   line_text);

        if (maxLineNumber <= line):
            funcNum = 1
            funcName = ""
            pos = 1
            pcall(15,fileNumber,fl,fn,funcNum,funcName,pos,line,maxLineNumber)

        fh.close()


    for col in range(1, 8):
        ws.cell(allLen, col).border = border_t

    wb.remove(wb['Sheet'])
    wb.save("test.xlsx")


if __name__ == "__main__":
    main()


